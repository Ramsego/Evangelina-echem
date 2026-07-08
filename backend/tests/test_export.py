"""Tests for the /export endpoint — generic plotted-only sheets + legacy etype path."""
import csv
import io


def test_xlsx_sheets_returns_named_workbook(client):
    """A sheets payload becomes a workbook with the given sheet, headers, units, rows."""
    import openpyxl
    r = client.post("/api/export", json={
        "format": "xlsx", "filename": "demo",
        "sheets": [{
            "name": "Plotted",
            "headers": ["Pot_1", "Cur_1"],
            "units":   ["V", "mA"],
            "rows":    [[0.10, 1.20], [0.20, 1.50]],
        }],
    })
    assert r.status_code == 200
    assert 'filename="demo.xlsx"' in r.headers["content-disposition"]
    wb = openpyxl.load_workbook(io.BytesIO(r.content))
    assert wb.sheetnames == ["Plotted"]
    assert wb.active.title == "Plotted"
    ws = wb.active
    assert [c.value for c in ws[1]] == ["Pot_1", "Cur_1"]   # headers
    assert [c.value for c in ws[2]] == ["V", "mA"]           # units
    assert ws["A3"].value == 0.10 and ws["B3"].value == 1.20  # first data row


def test_csv_sheets_has_headers_units_rows(client):
    r = client.post("/api/export", json={
        "format": "csv", "filename": "demo",
        "sheets": [{
            "name": "Plotted",
            "headers": ["Pot_1", "Cur_1"],
            "units":   ["V", "mA"],
            "rows":    [[0.1, 1.2]],
        }],
    })
    assert r.status_code == 200
    body = r.text
    assert "Pot_1,Cur_1" in body
    assert "V,mA" in body
    assert "0.1,1.2" in body


def test_export_requires_etype_when_no_sheets(client):
    """Without sheets and without etype the request is rejected (422), not a 500."""
    r = client.post("/api/export", json={"format": "xlsx", "filename": "demo"})
    assert r.status_code == 422


def test_csv_export_values_round_trip_exactly(client):
    """The numeric cells in a CSV export must equal what was plotted, byte for
    byte — this is the file an electrochemist opens in Excel/Origin and trusts.
    Parsed with Python's csv module (not substring matching) so quoting or
    delimiter issues would also be caught."""
    rows = [[0.1, 1.23456], [0.2, -2.5], [0.3, 0.0]]
    r = client.post("/api/export", json={
        "format": "csv", "filename": "roundtrip",
        "sheets": [{
            "name": "Plotted",
            "headers": ["Potential_V", "Current_mA"],
            "units":   ["V", "mA"],
            "rows":    rows,
        }],
    })
    assert r.status_code == 200
    parsed = list(csv.reader(io.StringIO(r.text)))
    assert parsed[0] == ["Potential_V", "Current_mA"]
    assert parsed[1] == ["V", "mA"]
    data_rows = [[float(c) for c in row] for row in parsed[2:]]
    assert data_rows == rows


def test_legacy_etype_curves_still_works(client, sine_curve):
    """Backward compatibility: the old etype/curves payload still builds a workbook."""
    import openpyxl
    r = client.post("/api/export", json={
        "format": "xlsx", "filename": "demo",
        "etype": "CV", "curves": [sine_curve],
        "scan_rate_mv": 10, "im_divisor": 1.0, "norm_label": "mA",
    })
    assert r.status_code == 200
    wb = openpyxl.load_workbook(io.BytesIO(r.content))
    assert "CV_Curves" in wb.sheetnames
