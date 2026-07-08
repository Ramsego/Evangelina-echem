import csv
import io
import os
import sys
import tempfile
from pathlib import Path
from typing import Literal

from fastapi import APIRouter, Request
from fastapi.responses import FileResponse, JSONResponse, StreamingResponse
from pydantic import BaseModel, Field, field_validator, model_validator

from limiter import limiter

router = APIRouter()

_MAX_PTS    = 20_000
_MAX_CURVES = 50


# ── Shared sub-models (mirrors analysis.py) ───────────────────────────────────

class Curve(BaseModel):
    vf: list[float]
    im: list[float]

    @field_validator("vf", "im")
    @classmethod
    def _check_len(cls, v: list[float]) -> list[float]:
        if len(v) > _MAX_PTS:
            raise ValueError(f"array exceeds {_MAX_PTS}-point limit")
        return v

    @model_validator(mode="after")
    def _check_match(self) -> "Curve":
        if len(self.vf) != len(self.im):
            raise ValueError("vf and im must have the same length")
        return self


class EISData(BaseModel):
    freq:  list[float]
    zreal: list[float]
    zimag: list[float]
    zmod:  list[float]
    zphz:  list[float]

    @field_validator("freq", "zreal", "zimag", "zmod", "zphz")
    @classmethod
    def _check_len(cls, v: list[float]) -> list[float]:
        if len(v) > _MAX_PTS:
            raise ValueError(f"array exceeds {_MAX_PTS}-point limit")
        return v


class GCDData(BaseModel):
    cycles:          list[int]
    discharge_caps:  list[float]
    charge_by_cycle: dict[str, float]
    dis_energy_mwh:  dict[str, float] | None = None
    ch_energy_mwh:   dict[str, float] | None = None


class Sheet(BaseModel):
    """A pre-built table the frontend hands over verbatim — the exact plotted
    columns, so no transform math is re-derived on the backend."""
    name:    str
    headers: list[str]
    units:   list[str] | None = None
    rows:    list[list[float | str | None]]


class ExportRequest(BaseModel):
    format:       Literal["xlsx", "csv", "opju"]
    etype:        Literal["CV", "LSV", "EISPOT", "GCD"] | None = None
    filename:     str                           = Field(max_length=200)
    sheets:       list[Sheet]   | None = None
    curves:       list[Curve]   | None = None
    scan_rate_mv: float         | None = None
    v_offset:     float         | None = None
    im_divisor:   float         | None = None
    norm_label:   str           | None = None   # e.g. "mA/cm²"
    eis:          EISData       | None = None
    gcd:          GCDData       | None = None

    @field_validator("curves")
    @classmethod
    def _check_curves(cls, v: list[Curve] | None) -> list[Curve] | None:
        if v is not None and len(v) > _MAX_CURVES:
            raise ValueError(f"exceeds {_MAX_CURVES}-curve limit")
        return v

    @model_validator(mode="after")
    def _require_etype_without_sheets(self) -> "ExportRequest":
        if self.sheets is None and self.etype is None:
            raise ValueError("etype is required when sheets is not provided")
        return self


# ── Origin detection ──────────────────────────────────────────────────────────

def _check_origin() -> tuple[bool, str | None]:
    if sys.platform != "win32":
        return False, "OriginPro is only available on Windows"
    origin_paths = [
        Path(os.environ.get("ProgramFiles",       "C:/Program Files"))       / "OriginLab",
        Path(os.environ.get("ProgramFiles(x86)",  "C:/Program Files (x86)")) / "OriginLab",
    ]
    if not any(p.exists() for p in origin_paths):
        return False, "OriginPro software not found on this machine"
    try:
        import originpro as op  # noqa: F401
        return True, None
    except ImportError:
        return False, "OriginPro detected — run `pip install originpro` in your backend environment to enable this"


# ── Capabilities endpoint ─────────────────────────────────────────────────────

@router.get("/capabilities")
@limiter.limit("30/minute")
def get_capabilities(request: Request):
    ok, msg = _check_origin()
    return {"xlsx": True, "csv": True, "opju": ok, "opju_message": msg}


# ── Main export endpoint ──────────────────────────────────────────────────────

@router.post("")
@limiter.limit("30/minute")
def export_data(request: Request, req: ExportRequest):
    stem = req.filename.replace(".dta", "").replace(".DTA", "")

    if req.format == "opju":
        ok, msg = _check_origin()
        if not ok:
            return JSONResponse(status_code=422, content={"detail": msg})
        return _export_opju(req, stem)

    if req.format == "xlsx":
        if req.sheets:
            return _export_xlsx_sheets(req.sheets, stem)
        return _export_xlsx(req, stem)

    if req.sheets:
        return _export_csv_sheets(req.sheets, stem)
    return _export_csv(req, stem)


# ── .opju builder (Windows + originpro installed) ─────────────────────────────

def _export_opju(req: ExportRequest, stem: str):
    import originpro as op

    tmp = tempfile.NamedTemporaryFile(suffix=".opju", delete=False)
    tmp.close()

    op.new()

    if req.sheets:
        # Plotted-only path: the frontend hands over the exact figure columns.
        # CV/LSV sheets are interleaved E/i pairs, so plot each odd column vs the
        # even column before it — mirrors the typed-curve graph loop below.
        sheet = req.sheets[0]
        wks   = op.new_sheet("w", lname=sheet.name[:31])
        cols  = list(zip(*sheet.rows)) if sheet.rows else [[] for _ in sheet.headers]
        for c, header in enumerate(sheet.headers):
            unit = sheet.units[c] if sheet.units and c < len(sheet.units) else ""
            wks.from_list(c, list(cols[c]) if c < len(cols) else [], lname=header, units=unit)

        gr    = op.new_graph(template="LINE")
        layer = gr[0]
        for c in range(1, len(sheet.headers), 2):
            layer.add_plot(wks, coly=c, colx=c - 1)
        layer.set_xlim(auto=True)
        gr.lname = f"{stem}"

        op.save(tmp.name)
        return FileResponse(
            tmp.name,
            media_type="application/octet-stream",
            filename=f"{stem}_origin.opju",
            background=_cleanup(tmp.name),
        )

    if req.etype in ("CV", "LSV") and req.curves:
        wks = op.new_sheet("w", lname="Curves")
        i_unit = req.norm_label or "mA"
        # Interleave E/i columns per curve
        for idx, curve in enumerate(req.curves):
            label = f"Curve {idx + 1}"
            e_col = idx * 2
            i_col = idx * 2 + 1
            wks.from_list(e_col, curve.vf,  lname=f"E_{idx+1}",  units="V")
            offset = req.v_offset or 0.0
            div    = req.im_divisor or 1.0
            scaled = [(v / div * 1000) for v in curve.im]  # A → mA (or mA/cm² etc.)
            wks.from_list(i_col, scaled, lname=f"i_{idx+1}", units=i_unit)
            wks.set_label(e_col, "Comment", label)

        # Settings sheet
        s = op.new_sheet("w", lname="Settings")
        s.from_list(0, ["Scan rate (mV/s)", "V offset (V)", "Normalization"], lname="Parameter")
        s.from_list(1, [
            str(req.scan_rate_mv or ""),
            str(req.v_offset or 0),
            req.norm_label or "none",
        ], lname="Value")

        # Graph
        gr    = op.new_graph(template="LINE")
        layer = gr[0]
        for idx in range(len(req.curves)):
            layer.add_plot(wks, coly=idx * 2 + 1, colx=idx * 2)
        layer.set_xlim(auto=True)
        layer.label("x", "E (V)")
        layer.label("y", req.norm_label or "Current (mA)")
        gr.lname = f"{stem}_CV"

    elif req.etype == "EISPOT" and req.eis:
        wks = op.new_sheet("w", lname="EIS_Data")
        wks.from_list(0, req.eis.freq,  lname="freq",  units="Hz")
        wks.from_list(1, req.eis.zreal, lname="Z_real", units="Ω")
        wks.from_list(2, req.eis.zimag, lname="Z_imag", units="Ω")
        wks.from_list(3, req.eis.zmod,  lname="Z_mod",  units="Ω")
        wks.from_list(4, req.eis.zphz,  lname="Phase",  units="deg")

        # Nyquist graph
        neg_zimag = [-z for z in req.eis.zimag]
        wks.from_list(5, neg_zimag, lname="-Z_imag", units="Ω")

        gr    = op.new_graph(template="SCATTER")
        layer = gr[0]
        layer.add_plot(wks, coly=5, colx=1)
        layer.label("x", "Z′ (Ω)")
        layer.label("y", "−Z″ (Ω)")
        gr.lname = f"{stem}_Nyquist"

    elif req.etype == "GCD" and req.gcd:
        gcd = req.gcd
        dis_mah = [c / 3600 * 1000 for c in gcd.discharge_caps]
        ce_vals: list[float | str] = []
        for cyc in gcd.cycles:
            ch = gcd.charge_by_cycle.get(str(cyc))
            dis = gcd.discharge_caps[gcd.cycles.index(cyc)] if cyc in gcd.cycles else None
            if ch and dis and ch > 0:
                ce_vals.append(round(dis / ch * 100, 2))
            else:
                ce_vals.append("")

        wks = op.new_sheet("w", lname="CycleLife")
        wks.from_list(0, gcd.cycles,   lname="Cycle",           units="")
        wks.from_list(1, dis_mah,       lname="Discharge_Cap",   units="mAh")
        wks.from_list(2, ce_vals,        lname="CE",              units="%")

        gr    = op.new_graph(template="SCATTER")
        layer = gr[0]
        layer.add_plot(wks, coly=1, colx=0)
        layer.label("x", "Cycle")
        layer.label("y", "Discharge Capacity (mAh)")
        gr.lname = f"{stem}_CycleLife"

    op.save(tmp.name)
    return FileResponse(
        tmp.name,
        media_type="application/octet-stream",
        filename=f"{stem}_origin.opju",
        background=_cleanup(tmp.name),
    )


# ── Generic sheets builders (plotted-only, columns handed over verbatim) ──────

def _export_xlsx_sheets(sheets: list[Sheet], stem: str):
    import openpyxl

    wb = openpyxl.Workbook()
    wb.remove(wb.active)  # remove default empty sheet
    for sheet in sheets:
        ws = wb.create_sheet(sheet.name[:31])  # Excel caps sheet names at 31 chars
        ws.append(sheet.headers)
        if sheet.units is not None:
            ws.append(sheet.units)
        for row in sheet.rows:
            ws.append(row)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{stem}.xlsx"'},
    )


def _export_csv_sheets(sheets: list[Sheet], stem: str):
    buf = io.StringIO()
    w   = csv.writer(buf)
    for i, sheet in enumerate(sheets):
        if i > 0:
            w.writerow([])
        if len(sheets) > 1:
            w.writerow([f"# === {sheet.name} ==="])
        w.writerow(sheet.headers)
        if sheet.units is not None:
            w.writerow(sheet.units)
        for row in sheet.rows:
            w.writerow(row)

    content = buf.getvalue().encode("utf-8")
    return StreamingResponse(
        io.BytesIO(content),
        media_type="text/csv",
        headers={"Content-Disposition": f'attachment; filename="{stem}.csv"'},
    )


# ── .xlsx builder ─────────────────────────────────────────────────────────────

def _export_xlsx(req: ExportRequest, stem: str):
    import openpyxl

    wb = openpyxl.Workbook()
    wb.remove(wb.active)  # remove default empty sheet

    if req.etype in ("CV", "LSV") and req.curves:
        ws = wb.create_sheet("CV_Curves")
        i_unit = req.norm_label or "mA"
        div    = req.im_divisor or 1.0
        offset = req.v_offset   or 0.0
        # Build interleaved columns: [E1, i1, E2, i2, ...]
        headers_name = []
        headers_unit = []
        columns: list[list[float]] = []
        for idx, curve in enumerate(req.curves):
            headers_name += [f"E_{idx+1}", f"i_{idx+1}"]
            headers_unit += ["V", i_unit]
            scaled = [(v / div * 1000) for v in curve.im]
            columns.append([v + offset for v in curve.vf])
            columns.append(scaled)
        ws.append(headers_name)
        ws.append(headers_unit)
        n_rows = max(len(c) for c in columns)
        for r in range(n_rows):
            ws.append([c[r] if r < len(c) else None for c in columns])

        ws2 = wb.create_sheet("CV_Settings")
        ws2.append(["Parameter", "Value"])
        ws2.append(["Scan rate (mV/s)", req.scan_rate_mv or ""])
        ws2.append(["V offset (V)", req.v_offset or 0])
        ws2.append(["Normalization", req.norm_label or "none"])

    elif req.etype == "EISPOT" and req.eis:
        ws = wb.create_sheet("EIS_Data")
        ws.append(["freq", "Z_real", "Z_imag", "Z_mod", "Phase"])
        ws.append(["Hz",   "Ω",      "Ω",      "Ω",     "deg"])
        for row in zip(req.eis.freq, req.eis.zreal, req.eis.zimag, req.eis.zmod, req.eis.zphz):
            ws.append(list(row))

    elif req.etype == "GCD" and req.gcd:
        gcd = req.gcd
        dis_mah = [c / 3600 * 1000 for c in gcd.discharge_caps]
        ce_vals: list[float | None] = []
        for cyc in gcd.cycles:
            ch  = gcd.charge_by_cycle.get(str(cyc))
            idx = gcd.cycles.index(cyc)
            dis = gcd.discharge_caps[idx]
            ce_vals.append(round(dis / ch * 100, 2) if ch and ch > 0 else None)

        ws = wb.create_sheet("GCD_CycleLife")
        ws.append(["Cycle", "Discharge_Cap", "CE"])
        ws.append(["",      "mAh",           "%"])
        for cyc, cap, ce in zip(gcd.cycles, dis_mah, ce_vals):
            ws.append([cyc, cap, ce])

        if gcd.dis_energy_mwh:
            we = wb.create_sheet("GCD_Energy")
            we.append(["Cycle", "Dis_Energy", "Ch_Energy", "Energy_Eff"])
            we.append(["",      "mWh",        "mWh",       "%"])
            for cyc in gcd.cycles:
                k   = str(cyc)
                dis = gcd.dis_energy_mwh.get(k)
                ch  = gcd.ch_energy_mwh.get(k) if gcd.ch_energy_mwh else None
                eff = round(dis / ch * 100, 2) if dis and ch and ch > 0 else None
                we.append([cyc, dis, ch, eff])

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{stem}_origin.xlsx"'},
    )


# ── .csv builder ──────────────────────────────────────────────────────────────

def _export_csv(req: ExportRequest, stem: str):
    buf  = io.StringIO()
    w    = csv.writer(buf)

    if req.etype in ("CV", "LSV") and req.curves:
        i_unit = req.norm_label or "mA"
        div    = req.im_divisor or 1.0
        offset = req.v_offset   or 0.0
        headers_name = []
        headers_unit = []
        columns: list[list[float]] = []
        for idx, curve in enumerate(req.curves):
            headers_name += [f"E_{idx+1}", f"i_{idx+1}"]
            headers_unit += ["V", i_unit]
            columns.append([v + offset for v in curve.vf])
            columns.append([(v / div * 1000) for v in curve.im])
        w.writerow(headers_name)
        w.writerow(headers_unit)
        n_rows = max(len(c) for c in columns)
        for r in range(n_rows):
            w.writerow([c[r] if r < len(c) else "" for c in columns])

    elif req.etype == "EISPOT" and req.eis:
        w.writerow(["freq", "Z_real", "Z_imag", "Z_mod", "Phase"])
        w.writerow(["Hz",   "Ω",      "Ω",      "Ω",     "deg"])
        for row in zip(req.eis.freq, req.eis.zreal, req.eis.zimag, req.eis.zmod, req.eis.zphz):
            w.writerow(list(row))

    elif req.etype == "GCD" and req.gcd:
        gcd = req.gcd
        dis_mah = [c / 3600 * 1000 for c in gcd.discharge_caps]
        w.writerow(["Cycle", "Discharge_Cap", "CE"])
        w.writerow(["",      "mAh",           "%"])
        for cyc, cap in zip(gcd.cycles, dis_mah):
            ch  = gcd.charge_by_cycle.get(str(cyc))
            ce  = round(cap / (ch / 3600 * 1000) * 100, 2) if ch and ch > 0 else ""
            w.writerow([cyc, cap, ce])

    content = buf.getvalue().encode("utf-8")
    return StreamingResponse(
        io.BytesIO(content),
        media_type="text/csv",
        headers={"Content-Disposition": f'attachment; filename="{stem}_origin.csv"'},
    )


# ── Cleanup helper ────────────────────────────────────────────────────────────

def _cleanup(path: str):
    from starlette.background import BackgroundTask
    def _rm():
        try:
            Path(path).unlink(missing_ok=True)
        except Exception:
            pass
    return BackgroundTask(_rm)
